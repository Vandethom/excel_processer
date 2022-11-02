from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('../wb/DI_DRIM_IdF.xlsx')
ws = wb.active

"""
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
"""

wb.create_sheet('Test')
wb.save('py_report.xlsx')

print(wb)
