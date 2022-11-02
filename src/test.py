import pandas as pd
from openpyxl import Workbook
from utils.excel_converter import WorkBook

new_dataframe = pd.DataFrame()

excel_export = WorkBook('../wb/DI_DRIM_IdF.xlsx', new_dataframe)

df_to_convert = pd.read_excel('../wb/DI_DRIM_IdF.xlsx')
excel_export.export_data()

"""
for row in range(1, 5):
    for col in range(1, 5):
        char = get_column_letter(col)
        cell = ws[char + str(row)]
        cell.value = None
        print(cell)
"""
"""
for cell in ws['A:A']:
    print(cell)
"""
